from leviosaPaperSubmission import app,paperSubUtils,paperSubUtilsFile
import sqlite3, copy, datetime, os, zipfile,shutil
from datetime import timedelta
from flask import request, session, g, redirect, url_for,\
	abort, render_template, flash, send_from_directory
from openpyxl import Workbook
from openpyxl.chart import Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.descriptors.excel import CellRange
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from shutil import copyfile
from openpyxl.utils import quote_sheetname
from openpyxl.styles import Border, Side




def generate_spreadsheet(paperInfoOfTheYear,tjyear):
	wb= Workbook()
	ws = wb.active
	ws.title = 'Submission Info'
	ws.cell(row = 1, column = 1, value = "Paper Id")
	ws.cell(row = 1, column = 2, value ="Title")
	ws.cell(row = 1, column = 3, value ="Author")
	ws.cell(row = 1, column = 4, value ="Department")
	ws.cell(row = 1, column = 5, value ="Manager")
	ws.cell(row = 1, column = 6, value ="Expert")
	ws.cell(row = 1, column = 7, value ="Manager Review Status")
	ws.cell(row = 1, column = 8, value ="Manager Review Result")
	ws.cell(row = 1, column = 9, value ="Expert Review Status")
	ws.cell(row = 1, column = 10, value ="Expert Review Result")


	i = 2
	for paper in paperInfoOfTheYear:
		managerReviewScore = 0
		for reviewItem in paper['managerReviewInfo']['itemList']:
			managerReviewScore += reviewItem['reviewItemGrade']
		expertReviewScore = 0
		for reviewItem in paper['expertReviewInfo']['itemList']:
			expertReviewScore += reviewItem['reviewItemGrade']
		if paper['managerReviewInfo']['reviewSubmitted'] == True:
			managerReviewStatus = "Submitted"
			if managerReviewScore < 10:
				score = '0'+str(managerReviewScore)
			else:
				score = str(managerReviewScore)
			managerReviewResult	 = 'Score:'+score+' Overall: '+app.config['REVIEW_RESULT'][paper['managerReviewInfo']['Overall']]
		else:
			managerReviewStatus = "Not submitted"
			managerReviewResult = "n/a"
		
		if paper['expertReviewInfo']['reviewSubmitted'] == True:
			expertReviewStatus = "Submitted"
			if expertReviewScore < 10:
				score = '0'+str(expertReviewScore)
			else:
				score = str(expertReviewScore)
			expertReviewResult	 = 'Score:'+score+' Overall: '+app.config['REVIEW_RESULT'][paper['expertReviewInfo']['Overall']]
		else:
			expertReviewStatus = "Not submitted"
			expertReviewResult = "n/a"

		ws.cell(row = i, column = 1, value = paper['paperInfo']['submissionSequence'])
		ws.cell(row = i, column = 2, value = paper['paperInfo']['paperTitle'])
		ws.cell(row = i, column = 3, value = paper['authorList'][0]['email'])
		ws.cell(row = i, column = 4, value = paper['paperInfo']['department'])
		ws.cell(row = i, column = 5, value = paper['managerInfo']['email'])
		ws.cell(row = i, column = 6, value = paper['expertInfo']['email'])
		ws.cell(row = i, column = 7, value = managerReviewStatus)
		ws.cell(row = i, column = 8, value = managerReviewResult)
		ws.cell(row = i, column = 9, value = expertReviewStatus)
		ws.cell(row = i, column = 10, value = expertReviewResult)
		i = i+1
	rangeTable = "A1:J%s" % (i-1)
	tab = Table(displayName="Table1", ref=rangeTable)

	# Add a default style with striped rows and banded columns
	style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
					   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
	tab.tableStyleInfo = style
	ws.add_table(tab)
	for column_cells in ws.columns:
		length = max(len(str(cell.value)) for cell in column_cells)
		ws.column_dimensions[column_cells[0].column].width = length
	
	ws = wb.create_sheet('files')
	ws.cell(row = 1, column = 1, value = "Paper Id")
	ws.cell(row = 1, column = 2, value ="Author")
	ws.cell(row = 1, column = 3, value ="PDF link")
	ws.cell(row = 1, column = 4, value ="PDF")
	ws.cell(row = 1, column = 5, value ="ZIP link")
	ws.cell(row = 1, column = 6, value ="ZIP")
	ws.cell(row = 1, column = 7, value ="Manager excel link")
	ws.cell(row = 1, column = 8, value ="Manager excel")
	ws.cell(row = 1, column = 9, value ="Expert excel link")
	ws.cell(row = 1, column = 10, value ="Expert excel")
	ws.cell(row = 1, column = 11, value ="Manager PDF link")
	ws.cell(row = 1, column = 12, value ="Manager PDF")
	ws.cell(row = 1, column = 13, value ="Expert PDF link")
	ws.cell(row = 1, column = 14, value ="Expert PDF")

	i = 2
	for paper in paperInfoOfTheYear:
		
		ws.cell(row = i, column = 1, value = paper['paperInfo']['submissionSequence'])
		ws.cell(row = i, column = 2, value = paper['authorList'][0]['email'])
		filenameStr = paperSubUtilsFile.form_file_name(tjyear,paper['paperInfo']['submissionSequence'],paper['paperInfo'])
		
		j = 3
		for fileType in ['PDF','ZIP']:
			if fileType == 'PDF': 
				filename = filenameStr+'.pdf'
				urlDisplay = 'pdfURL'
			elif fileType == 'ZIP': 
				filename = filenameStr+'.zip'
				urlDisplay = 'zipURL'
		
			file = os.path.join(app.config['UPLOAD_FOLDER'],filename)
			if(os.path.isfile(file)):
				ws.cell(row = i, column = j, value = "yes")
			else:
				ws.cell(row = i, column = j, value = "no")
			ws.cell(row = i, column = j+1, value = paper['paperInfo'][urlDisplay])
			j = j+2

		for fileType in ['ManagerEXCEL','ExpertEXCEL','ManagerPDF','ExpertPDF']:
			if fileType == 'ManagerEXCEL': 
				filename = filenameStr+'manager_Review'+'.xlsx'
				reviewInfo = 'managerReviewInfo'
				urlDisplay = 'excelURL'
			elif fileType == 'ExpertEXCEL':				
				filename = filenameStr+'expert_Review'+'.xlsx'
				reviewInfo = 'expertReviewInfo'
				urlDisplay = 'excelURL'
			elif fileType == 'ManagerPDF': 
				filename = filenameStr+'manager_Review'+'.pdf'
				reviewInfo = 'managerReviewInfo'
				urlDisplay = 'commentedVersionURL'
			elif fileType == 'ExpertPDF': 
				filename = filenameStr+'expert_Review'+'.pdf'
				reviewInfo = 'expertReviewInfo'
				urlDisplay = 'commentedVersionURL'
		
			file = os.path.join(app.config['UPLOAD_FOLDER'],filename)
			if(os.path.isfile(file)):
				ws.cell(row = i, column = j, value = "yes")
			else:
				ws.cell(row = i, column = j, value = "no")
			ws.cell(row = i, column = j+1, value = paper[reviewInfo][urlDisplay])
			j = j+2
 
		i = i+1
	rangeTable = "A1:N%s" % (i-1)
	tab = Table(displayName="Table1", ref=rangeTable)

	# Add a default style with striped rows and banded columns
	style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
					   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
	tab.tableStyleInfo = style
	ws.add_table(tab)
	for column_cells in ws.columns:
		length = max(len(str(cell.value)) for cell in column_cells)
		ws.column_dimensions[column_cells[0].column].width = length
	
	wb.save(os.path.join(app.config['UPLOAD_FOLDER'],str(tjyear)+'.xlsx'))
	paperSubUtils.database_operation_update_tjinfo(tjyear, 'spreadsheet')
	return


def download_archive(paperInfoOfTheYear,tjyear):
	downloadZipFile = zipfile.ZipFile(os.path.join(app.config['UPLOAD_FOLDER'],str(tjyear)+'.zip'),'w')
	for file in os.listdir(app.config['UPLOAD_FOLDER']):
		if file.startswith(str(tjyear)+'-'):
			downloadZipFile.write(os.path.join(app.config['UPLOAD_FOLDER'],file),\
					file,\
					compress_type = zipfile.ZIP_DEFLATED)
	downloadZipFile.close()

	paperSubUtils.database_operation_update_tjinfo(tjyear, 'archive')
	return

def generate_review_form_spreadsheet(paperInfoWithReview, tjyear, isManager, isPaperSub = False):
	filenameStr = paperSubUtilsFile.form_file_name(tjyear,\
		paperInfoWithReview['paperInfo']['submissionSequence'],paperInfoWithReview['paperInfo'])



	if isManager == True:
		filenameStr = filenameStr+'manager_Review'
		if isPaperSub == False:
			reviewInfo = paperInfoWithReview['managerReviewInfo']
	else:
		filenameStr = filenameStr+'expert_Review'
		if isPaperSub == False:
			reviewInfo = paperInfoWithReview['expertReviewInfo']
	filename = filenameStr+'.xlsx'
	copyfile(os.path.join(app.config['UPLOAD_FOLDER'],'review_form.xlsx'), os.path.join(app.config['UPLOAD_FOLDER'],filename))
	wb = load_workbook(os.path.join(app.config['UPLOAD_FOLDER'],filename))

	ws = wb.active
	ws.cell(row = 1, column = 1, value = app.config['EDITORIAL_COMMITEE_INFO']['journal']+ app.config['EDITORIAL_COMMITEE_INFO']['form'])
	ws.cell(row = 3, column = 2, value = paperInfoWithReview['paperInfo']['paperTitle'])
	ws.cell(row = 4, column = 2, value = paperInfoWithReview['authorList'][0]['email'])
	if isManager == True:
		ws.cell(row = 5, column = 2, value = paperInfoWithReview['managerInfo']['email'])
	else:
		ws.cell(row = 5, column = 2, value = paperInfoWithReview['expertInfo']['email'])

	i=0
	for reviewItemName in app.config['REVIEW_ITEM']:
		ws.cell(row=9+i,column=1,value=reviewItemName)
		i = i+1

	i = 0
	for question in app.config['REVIEW_COMMENTS_QUESTIONS']:
		ws.cell(row=19+i*5, column= 1, value = question)
		i = i+1
		
	set_border(ws, 'A18:D43')
	if isPaperSub == False:
		if reviewInfo['reviewSubmitted']==True:
			i=0
			for reviewItem in reviewInfo['itemList']:
				ws.cell(row=9+i,column=3,value=app.config['REVIEW_RESULT'][reviewItem['reviewItemGrade']])
				i = i+1
			ws.cell(row=16, column=3, value=app.config['REVIEW_RESULT'][reviewInfo['Overall']])
			i = 0
			for reviewQuestion in reviewInfo['questionList']:
				ws.cell(row=20+i*5, column= 1, value = reviewQuestion['answer'])
				i = i+1
		else:
			ws_bk = sheet = wb.get_sheet_by_name('bk')
			i=1
			string = ""
			for result in app.config['REVIEW_RESULT']:
				ws_bk.cell(row=1,column=i, value=result)
				i = i+1
				
			
			dv = DataValidation(type="list",\
							formula1="{0}!$A$1:$D$1".format(quote_sheetname('bk'))\
							)

								


			#Optionally set a custom error message
			dv.error ='Your entry is not in the list'
			dv.errorTitle = 'Invalid Entry'
			
			#Optionally set a custom prompt message
			dv.prompt = 'Please select from the list'
			dv.promptTitle = 'List Selection'


			ws.add_data_validation(dv)
			dv.add('C9:c16')
	else:
		ws_bk = sheet = wb.get_sheet_by_name('bk')
		i=1
		string = ""
		for result in app.config['REVIEW_RESULT']:
			ws_bk.cell(row=1,column=i, value=result)
			i = i+1
			
		
		dv = DataValidation(type="list",\
						formula1="{0}!$A$1:$D$1".format(quote_sheetname('bk'))\
						)



		#Optionally set a custom error message
		dv.error ='Your entry is not in the list'
		dv.errorTitle = 'Invalid Entry'
		
		#Optionally set a custom prompt message
		dv.prompt = 'Please select from the list'
		dv.promptTitle = 'List Selection'


		ws.add_data_validation(dv)
		dv.add('C9:c16')
		
	wb.save(os.path.join(app.config['UPLOAD_FOLDER'],filename))
	
	
	return filename

def set_border(ws, cell_range):
	rows = ws[cell_range]
	side = Side(border_style='thin', color="FF000000")

	rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
	max_y = len(rows) - 1  # index of the last row
	for pos_y, cells in enumerate(rows):
		max_x = len(cells) - 1	# index of the last cell
		for pos_x, cell in enumerate(cells):
			border = Border(
				left=cell.border.left,
				right=cell.border.right,
				top=cell.border.top,
				bottom=cell.border.bottom
			)
			if pos_x == 0:
				border.left = side
			if pos_x == max_x:
				border.right = side
			border.top = side
			border.bottom = side

			cell.border = border
			
def extract_info_from_review_spreadsheet(paperInfoWithReview, tjyear, isManager, reviewInfo):
	filenameStr = paperSubUtilsFile.form_file_name(tjyear,\
		paperInfoWithReview['paperInfo']['submissionSequence'],paperInfoWithReview['paperInfo'])

	if isManager == True:
		filenameStr = filenameStr+'manager_Review'

	else:
		filenameStr = filenameStr+'expert_Review'

	filename = filenameStr+'.xlsx'
	wb = load_workbook(os.path.join(app.config['UPLOAD_FOLDER'],filename))
	ws = wb.active
	
	rowIndex=0
	for eachReviewItem in reviewInfo['itemList']:
		eachReviewItem['reviewItemGrade'] =app.config['REVIEW_RESULT'].index(ws.cell(row=9+rowIndex,column=3).value)
		rowIndex = rowIndex+1

	reviewInfo['Overall'] = app.config['REVIEW_RESULT'].index(ws.cell(row=16,column=3).value)
 	rowIndex = 0
	for eachReviewQuestion in reviewInfo['questionList']:
		if ws.cell(row=20+rowIndex*5, column= 1).value != None:
			eachReviewQuestion['answer']=ws.cell(row=20+rowIndex*5, column= 1).value
		else:
			eachReviewQuestion['answer']="n/a"
		rowIndex = rowIndex+1
	return reviewInfo
